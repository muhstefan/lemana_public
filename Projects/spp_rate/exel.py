from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from time_module import SppRateTime


class ExelEngine:

    def __init__(self, current_world):
        self.Excel_name = f"Данные - {SppRateTime.yesterday}.xlsx"
        self.book = None
        self.sheet = None
        self.spp = None
        self.spp_list = None
        self.current_world = current_world

    color_dict = {
        "coral": "FF7F50",
        "salad": "90EE90",
        "pale_yellow": "FFFFE0",
        "aquamarine": "4FFFC5",
        "peach": "FFB69C",
        "lemon chiffon": "FFF5B2",
        "coral2": "FFC0CB",
    }

    def set_spp_list(self, spp_list):
        self.spp_list = spp_list

    def activate(self):
        self._add_date_row()
        for spp in self.spp_list.values():
            self.spp = spp
            self._add_exel_data()
            self._exel_format_color()
            self._exel_format_style()
        self._exel_format_width()
        self.book.save(self.Excel_name)

    def excel_check(self):
        try:
            # Пытаемся загрузить файл
            self.book = load_workbook(self.Excel_name)
            self.sheet = self.book.active

        except FileNotFoundError:
            # Если файл не существует, создаем его
            self._init_book()

    def _init_book(self):
        self.book = Workbook()
        self.sheet = self.book.active

        # Заполняем заголовки в первой строке
        headers = list(
            self.current_world.ExData.keys()
        )  # Получаем ключи из хэш-таблицы
        self.sheet.append(headers)  # Добавляем заголовки в первую строку
        # Сохраняем книгу
        self.book.save(self.Excel_name)

    def _color(self, color_name):
        hex_code = self.color_dict.get(color_name)

        if hex_code:
            return PatternFill(
                start_color=hex_code,
                end_color=hex_code,
                fill_type="solid",  # Можно жестко задать 'solid', если не нужны другие типы
            )
        return None

    def _add_date_row(self):

        next_row = self.sheet.max_row + 1

        # Настройки стиля
        font = Font(size=14, bold=True, name="Calibri Light")
        alignment = Alignment(horizontal="center", vertical="center")
        fill = self._color("coral2")

        # Записываем текст в первую ячейку
        cell = self.sheet.cell(row=next_row, column=1, value=SppRateTime.yesterday)
        cell.font = font
        cell.alignment = alignment
        cell.fill = fill

        for col in range(1, self.sheet.max_column + 1):
            self.sheet.cell(row=next_row, column=col).fill = fill

    def _add_exel_data(self):
        temp_table = self.current_world.ExData
        temp_table["ИМЯ"] = self.spp.name
        temp_table["Points"] = self.spp.points
        for key in self.spp.actions_stat:
            temp_table[key] = self.spp.actions_stat[key]
        for i in range(24):
            temp_table[i] = self._exel_time_stat(
                self.spp, i
            )  # Заполняем элементы от time_stat

        self._exel_spp_colors()
        self._add_data(temp_table)

    @staticmethod
    def _exel_time_stat(spp, entry_index):
        # Получаем нужную запись (переводим 1-based индекс в 0-based)
        time_stat_entry = spp.time_stat[entry_index]

        formatted_lines = []
        for parameter_name, value in time_stat_entry.items():
            if value != 0:
                # Разбиваем название параметра на слова
                words = parameter_name.split()
                # Если два слова - берем первые буквы каждого
                if len(words) == 2:
                    first_letters = words[0][0] + words[1][0]
                else:
                    # Иначе берем только первую букву первого слова
                    first_letters = words[0][0]
                formatted_lines.append(f"{first_letters}: {value}")

        return "\n".join(formatted_lines)

    def _exel_spp_colors(self):
        self.spp.color = "salad"

    def _add_data(self, data):

        next_row = self.sheet.max_row + 1
        # Проходим по каждому ключу в данных
        for column_index, (key, value) in enumerate(data.items(), start=1):
            value = data[key]  # Получаем значение для текущего ключа
            self.sheet.cell(row=next_row, column=column_index, value=value)

    def _exel_format_style(self):

        current_row = self.sheet.max_row
        thin = Side(border_style="thin", color="000000")  # Черная обводка
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        indices = range(len(self.current_world.ExData))
        for key in indices:
            cell = self.sheet.cell(row=current_row, column=key + 1)
            cell.alignment = Alignment(
                wrap_text=True, horizontal="center", vertical="center"
            )  # переносы по словам
            cell.font = Font(bold=True, size=10)
            cell.border = border

        for i, key in enumerate(self.spp.actions_stat):
            cell = self.sheet.cell(row=current_row, column=i + 3)
            cell.font = Font(size=20, name="Calibri Light")

        cell = self.sheet.cell(row=current_row, column=1)
        cell.font = Font(size=16, name="Calibri Light")
        cell = self.sheet.cell(row=current_row, column=2)
        cell.font = Font(size=20, name="Calibri Light")

    def _exel_format_width(self):

        # Устанавливаем ширину столбцов
        counter = 3
        start_format = 3
        self.sheet.column_dimensions["A"].width = 25
        self.sheet.column_dimensions["B"].width = 16
        for i, key in enumerate(self.spp.actions_stat):
            self.sheet.column_dimensions[get_column_letter(i + start_format)].width = 16
            counter += 1
        for i, key in enumerate(self.spp.time_stat):  # C = 2, T = 19 (индексы)
            self.sheet.column_dimensions[get_column_letter(i + counter)].width = 6

        # Форматируем заголовки
        for cell in self.sheet[1]:  # Первая строка (заголовки)
            cell.font = Font(bold=True, color="FFFFFF")  # Жирный шрифт, белый цвет
            cell.fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )  # Синий фон
            cell.alignment = Alignment(
                horizontal="center", vertical="center"
            )  # Центрируем текст

    def _exel_format_color(self):

        current_row = self.sheet.max_row
        self._use_color_rule("salad", "coral", "70", "Points")
        self._use_color_rule("salad", "pale_yellow", "3", "Отрисовки")
        self._use_color_rule("salad", "pale_yellow", "0", "Продажи")
        self._use_color_rule("salad", "pale_yellow", "0", "Услуги")
        self._use_color_rule("coral", "pale_yellow", "0", "Не закрыл")
        self._use_color_rule("coral", "pale_yellow", "0", "Поздно взял")
        # .fill это не метод, а свойство (поле) которому мы присваиваем цвет
        self.sheet.cell(row=current_row, column=1).fill = self._color(self.spp.color)

        # т.к у нас есть еще какие-то колонны тут у нас поправка
        column_delta = (
            len(self.current_world.names) + len(self.current_world.spp_rate_logic) + 1
        )
        day_begin = self.spp.day_begin + column_delta
        day_end = self.spp.day_end + column_delta

        for col in range(day_begin, day_end):
            cell = self.sheet.cell(row=current_row, column=col)
            cell.fill = self._color(self.spp.color)  # Применяем цветовое заполнение
        self.book.save(self.Excel_name)

    def _find_column(self, header_name):
        for cell in self.sheet[1]:  # Ищем в первой строке
            if cell.value == header_name:
                return get_column_letter(cell.column)
        return None

    def _use_color_rule(self, color_1, color_2, digit_as_string, header_name):

        column = self._find_column(header_name)
        if column:
            cur_row = self.sheet.max_row
            init_color_1 = self._color(color_1)
            init_color_2 = self._color(color_2)
            init_rule_greater = CellIsRule(
                operator="greaterThan", formula=[digit_as_string], fill=init_color_1
            )
            init_rule_less = CellIsRule(
                operator="lessThan", formula=[digit_as_string], fill=init_color_2
            )
            self.sheet.conditional_formatting.add(
                f"{column}{cur_row}", init_rule_greater
            )
            self.sheet.conditional_formatting.add(f"{column}{cur_row}", init_rule_less)
